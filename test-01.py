import ctypes
import base64
import openpyxl
from openpyxl import Workbook
import os


# 定义结构体（同之前，单个结构体大小1112字节）
class STExportValue_t(ctypes.Structure):
    _fields_ = [
        ("data_len", ctypes.c_int),
        ("data_buf", ctypes.c_char * 1024),  # 512 * 2=1024
        ("time_len", ctypes.c_int),
        ("time_buf", ctypes.c_char * 32),
        ("protocol_len", ctypes.c_int),
        ("protocol_type_buf", ctypes.c_char * 7),
        ("analysis_len", ctypes.c_int),
        ("analysis_result_buf", ctypes.c_char * 32),
    ]


# 单个结构体的字节大小（验证是否为1112）
STRUCT_SIZE = ctypes.sizeof(STExportValue_t)
assert STRUCT_SIZE == 1112, f"结构体大小错误，实际为{STRUCT_SIZE}字节"


def parse_multiple_structs(binary_data):
    """解析多组结构体数据，返回列表"""
    struct_list = []
    data_length = len(binary_data)

    # 检查总长度是否为结构体大小的整数倍（避免数据不完整）
    if data_length % STRUCT_SIZE != 0:
        raise ValueError(f"二进制数据长度不正确，应为{STRUCT_SIZE}的整数倍，实际为{data_length}")

    # 按结构体大小分割并解析
    for i in range(0, data_length, STRUCT_SIZE):
        # 截取当前结构体的二进制数据
        chunk = binary_data[i:i + STRUCT_SIZE]
        # 转换为结构体实例
        struct_instance = STExportValue_t.from_buffer_copy(chunk)
        struct_list.append(struct_instance)

    return struct_list


def create_excel_template():
    """创建Excel模板，前2行保持不变"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 第一行：标题
    ws['A1'] = "馈线终端离线分析装置-导出数据表"

    # 第二行：表头
    headers = ["时间", "协议类型", "源数据", "分析结果"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)

    return wb, ws


def save_to_excel(structs, output_filename="导出数据表.xlsx"):
    """将解析的数据保存到Excel文件"""

    # 创建或加载模板
    if os.path.exists(output_filename):
        wb = openpyxl.load_workbook(output_filename)
        ws = wb.active
        # 清空第3行及以后的数据（保留前2行）
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)
    else:
        wb, ws = create_excel_template()

    # 从第3行开始写入数据
    start_row = 3

    for idx, uni in enumerate(structs):
        row = start_row + idx

        # 处理时间字段
        time_buf = uni.time_buf.decode('utf-8', errors='ignore')
        time_buf = " ".join([item.strip() for item in time_buf.strip().split("\n")])

        # 处理协议类型字段
        protocol_type_buf = uni.protocol_type_buf.decode('utf-8', errors='ignore')

        # 处理源数据字段（Base64解码）
        data_buf_base64_str = uni.data_buf.decode('utf-8', errors='ignore')
        try:
            decoded_bytes = base64.b64decode(data_buf_base64_str.encode('utf-8'))
            data_buf = decoded_bytes.decode('utf-8', errors='ignore')
        except Exception as e:
            data_buf = f"解码错误: {str(e)}"

        # 处理分析结果字段
        analysis_result_buf = uni.analysis_result_buf.decode('utf-8', errors='ignore')

        # 写入Excel
        ws.cell(row=row, column=1, value=time_buf)  # A列：时间
        ws.cell(row=row, column=2, value=protocol_type_buf)  # B列：协议类型
        ws.cell(row=row, column=3, value=data_buf)  # C列：源数据
        ws.cell(row=row, column=4, value=analysis_result_buf)  # D列：分析结果

    # 保存文件
    wb.save(output_filename)
    print(f"数据已成功导出到 {output_filename}")


# ------------------------------
# 使用示例：解析包含多组数据的二进制文件并导出到Excel
# ------------------------------
if __name__ == "__main__":
    # 读取二进制文件（假设包含多组结构体数据）
    with open("2025-10-28-10-04-09.txt", "rb") as f:
        all_binary_data = f.read()

    try:
        # 解析所有数据
        structs = parse_multiple_structs(all_binary_data)
        print(f"成功解析 {len(structs)} 组数据")

        # 将数据保存到Excel
        save_to_excel(structs)

        # 可选：在控制台显示解析结果（调试用）
        if structs:
            for idx, uni in enumerate(structs):
                print("====================== 第{}组数据: Start ============".format(idx + 1))
                print("data_len:", uni.data_len)

                data_buf_base64_str = uni.data_buf.decode('utf-8', errors='ignore')
                try:
                    decoded_bytes = base64.b64decode(data_buf_base64_str.encode('utf-8'))
                    decoded_str = decoded_bytes.decode('utf-8', errors='ignore')
                    print("data_buf:", decoded_str)
                except Exception as e:
                    print("data_buf: 解码错误 -", str(e))

                print("time_len:", uni.time_len)
                time_buf = uni.time_buf.decode('utf-8', errors='ignore')
                time_buf = " ".join([item.strip() for item in time_buf.strip().split("\n")])
                print("time_buf:", time_buf)

                print("protocol_len:", uni.protocol_len)
                protocol_type_buf = uni.protocol_type_buf.decode('utf-8', errors='ignore')
                print("protocol_type_buf:", protocol_type_buf)

                print("analysis_len:", uni.analysis_len)
                analysis_result_buf = uni.analysis_result_buf.decode('utf-8', errors='ignore')
                print("analysis_result_buf:", analysis_result_buf)
                print("====================== 第{}组数据: END ============".format(idx + 1))

    except ValueError as e:
        print(f"解析失败：{e}")
    except Exception as e:
        print(f"处理过程中发生错误：{e}")