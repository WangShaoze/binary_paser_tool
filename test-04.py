import ctypes
import base64
import openpyxl
import os
from datetime import datetime
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, \
    QCheckBox, QLineEdit, QGroupBox, QFormLayout, QLabel, QStatusBar, QDialog, QMessageBox, QScrollArea
from PyQt5.QtCore import Qt, QSettings
from PyQt5.QtGui import QIcon
import sys


# 定义结构体（同之前，单个结构体大小1112字节）
class STExportValue_t(ctypes.Structure):
    _fields_ = [
        ("data_len", ctypes.c_int),
        ("data_buf", ctypes.c_char * 1024),  # 512 * 2 = 1024
        ("time_len", ctypes.c_int),
        ("time_buf", ctypes.c_char * 32),
        ("protocol_len", ctypes.c_int),
        ("protocol_type_buf", ctypes.c_char * 7),
        ("analysis_len", ctypes.c_int),
        ("analysis_result_buf", ctypes.c_char * 32),
    ]


# 单个结构体的字节大小（验证是否为1112）
STRUCT_SIZE = ctypes.sizeof(STExportValue_t)
assert STRUCT_SIZE == 1112, f"结构体大小错误，实际为{STRUCT_SIZE}字节"


def parse_multiple_structs(binary_data):
    """解析多组结构体数据，返回列表"""
    struct_list = []
    data_length = len(binary_data)

    # 检查总长度是否为结构体大小的整数倍（避免数据不完整）
    if data_length % STRUCT_SIZE != 0:
        raise ValueError(f"二进制数据长度不正确，应为{STRUCT_SIZE}的整数倍，实际为{data_length}")

    # 按结构体大小分割并解析
    for i in range(0, data_length, STRUCT_SIZE):
        # 截取当前结构体的二进制数据
        chunk = binary_data[i:i + STRUCT_SIZE]
        # 转换为结构体实例
        struct_instance = STExportValue_t.from_buffer_copy(chunk)
        struct_list.append(struct_instance)

    return struct_list


def create_excel_template():
    """创建Excel模板，前2行保持不变"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 第一行：标题 - 合并ABCD四列
    ws.merge_cells('A1:D1')
    ws['A1'] = "馈线终端离线分析装置-导出数据表"

    # 设置合并单元格的样式使其居中显示
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 第二行：表头
    headers = ["时间", "协议类型", "源数据", "分析结果"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)

    return wb, ws


def generate_timestamp_filename():
    """生成包含当前日期时间的文件名"""
    now = datetime.now()
    timestamp = now.strftime("%Y年%m月%d日%H时%M分%S秒")
    return f"导出数据表-{timestamp}.xlsx"


def save_to_excel(structs, output_filename=None):
    """将解析的数据保存到Excel文件"""
    # 如果没有提供文件名，使用当前日期时间生成
    if output_filename is None:
        output_filename = generate_timestamp_filename()

    # 如果文件已存在，则加载文件
    if os.path.exists(output_filename):
        wb = openpyxl.load_workbook(output_filename)
        ws = wb.active
        # 找到最后一行的下一行作为开始写入的行
        start_row = ws.max_row + 1
    else:
        # 如果文件不存在，则创建一个新的文件并写入表头
        wb, ws = create_excel_template()
        start_row = 3  # 从第三行开始写入数据，因为前两行是表头

    # 从指定的起始行开始写入数据
    for idx, uni in enumerate(structs):
        row = start_row + idx

        # 处理时间字段
        time_buf = uni.time_buf.decode('utf-8', errors='ignore')
        time_buf = " ".join([item.strip() for item in time_buf.strip().split("\n")])

        # 处理协议类型字段
        protocol_type_buf = uni.protocol_type_buf.decode('utf-8', errors='ignore')

        # 处理源数据字段（Base64解码）
        data_buf_base64_str = uni.data_buf.decode('utf-8', errors='ignore')
        try:
            decoded_bytes = base64.b64decode(data_buf_base64_str.encode('utf-8'))
            data_buf = decoded_bytes.decode('utf-8', errors='ignore')
        except Exception as e:
            data_buf = f"解码错误: {str(e)}"

        # 处理分析结果字段
        analysis_result_buf = uni.analysis_result_buf.decode('utf-8', errors='ignore')

        # 写入Excel
        ws.cell(row=row, column=1, value=time_buf)  # A列：时间
        ws.cell(row=row, column=2, value=protocol_type_buf)  # B列：协议类型
        ws.cell(row=row, column=3, value=data_buf)  # C列：源数据
        ws.cell(row=row, column=4, value=analysis_result_buf)  # D列：分析结果

    # 保存文件
    wb.save(output_filename)
    print(f"数据已成功导出到 {output_filename}")


def parse_file(file_path):
    """解析指定的二进制文件并导出到Excel"""
    try:
        # 读取二进制文件
        with open(file_path, "rb") as f:
            all_binary_data = f.read()

        # 解析所有数据
        structs = parse_multiple_structs(all_binary_data)
        print(f"成功解析 {len(structs)} 组数据")

        return structs
    except ValueError as e:
        print(f"解析失败：{e}")
    except Exception as e:
        print(f"处理过程中发生错误：{e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("二进制数据解析工具")
        self.setGeometry(100, 100, 800, 600)

        # 设置窗口图标
        icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icon", "app.ico")
        self.setWindowIcon(QIcon(icon_path))

        self.file_path = None
        self.output_folder = None
        self.selected_files = []

        # 加载上次选择的路径
        self.settings = QSettings("MyCompany", "BinaryParserApp")
        self.last_used_folder = self.settings.value("last_used_folder", "", type=str)
        self.last_output_folder = self.settings.value("last_output_folder", "", type=str)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setSpacing(15)  # 设置整体布局间距
        layout.setContentsMargins(20, 20, 20, 20)  # 设置边距

        # 文件路径选择区域
        file_group = QGroupBox("选择文件路径")
        file_layout = QFormLayout()
        file_layout.setSpacing(12)  # 设置表单布局间距
        file_layout.setContentsMargins(15, 15, 15, 15)  # 设置表单边距

        # 数据文件路径行（输入框+按钮）
        data_file_row = QHBoxLayout()
        self.data_file_path_edit = QLineEdit(self)
        self.data_file_path_edit.setReadOnly(True)
        self.data_file_path_edit.setFixedHeight(35)  # 调整文件路径框的高度
        data_file_button = QPushButton("浏览", self)
        data_file_button.setFixedHeight(35)
        data_file_button.setFixedWidth(80)
        data_file_button.clicked.connect(self.select_data_file_path)
        data_file_row.addWidget(self.data_file_path_edit)
        data_file_row.addWidget(data_file_button)
        data_file_row.setSpacing(10)
        file_layout.addRow("数据文件路径:", data_file_row)

        # 保存路径行（输入框+按钮）
        save_path_row = QHBoxLayout()
        self.save_path_edit = QLineEdit(self)
        self.save_path_edit.setReadOnly(True)
        self.save_path_edit.setFixedHeight(35)
        save_button = QPushButton("浏览", self)
        save_button.setFixedHeight(35)
        save_button.setFixedWidth(80)
        save_button.clicked.connect(self.select_save_path)
        save_path_row.addWidget(self.save_path_edit)
        save_path_row.addWidget(save_button)
        save_path_row.setSpacing(10)
        file_layout.addRow("保存路径:", save_path_row)

        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # 刷新按钮
        refresh_button = QPushButton("刷新", self)
        refresh_button.setFixedHeight(40)
        refresh_button.clicked.connect(self.refresh)
        layout.addWidget(refresh_button)

        # 全选和反选按钮
        select_all_button = QPushButton("全选", self)
        select_all_button.setFixedHeight(40)
        select_all_button.clicked.connect(self.select_all_files)

        deselect_all_button = QPushButton("反选", self)
        deselect_all_button.setFixedHeight(40)
        deselect_all_button.clicked.connect(self.deselect_all_files)

        file_selection_layout = QHBoxLayout()
        file_selection_layout.setSpacing(10)
        file_selection_layout.addWidget(select_all_button)
        file_selection_layout.addWidget(deselect_all_button)
        layout.addLayout(file_selection_layout)

        # 文件选择区域
        self.files_group = QGroupBox("选择需要解析的文件")
        self.files_group.setEnabled(False)  # 默认禁用

        # 为文件选择区域添加滚动功能
        self.files_scroll_area = QWidget()
        self.files_layout = QVBoxLayout(self.files_scroll_area)
        self.files_layout.setSpacing(8)

        scroll = QScrollArea()
        scroll.setWidget(self.files_scroll_area)
        scroll.setWidgetResizable(True)
        scroll.setFixedHeight(200)

        files_inner_layout = QVBoxLayout()
        files_inner_layout.addWidget(scroll)
        self.files_group.setLayout(files_inner_layout)

        layout.addWidget(self.files_group)

        # 状态栏
        self.status_bar = QStatusBar(self)
        self.setStatusBar(self.status_bar)

        # 开始解析按钮
        parse_button = QPushButton("开始解析", self)
        parse_button.setFixedHeight(50)
        parse_button.clicked.connect(self.start_parsing)
        layout.addWidget(parse_button)

        # 设置主界面布局
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # 设置样式
        self.set_style()

        # 显示上次保存的路径
        if self.last_used_folder:
            self.data_file_path_edit.setText(self.last_used_folder)
            self.file_path = self.last_used_folder
            self.load_files(self.last_used_folder)
        if self.last_output_folder:
            self.save_path_edit.setText(self.last_output_folder)
            self.output_folder = self.last_output_folder

        # 初始化状态栏
        self.update_status("ready")

    def set_style(self):
        """设置界面样式"""
        self.setStyleSheet("""
            /* 窗口背景 */
            QMainWindow {
                background-color: #f5f5f5;
            }
            
            /* 分组框样式 */
            QGroupBox {
                font: bold 14pt "微软雅黑";
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                margin-top: 20px;
                background-color: white;
            }
            
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 15px;
                background-color: #f5f5f5;
                color: #333;
            }
            
            /* 按钮样式 */
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-family: "微软雅黑";
                border: none;
                border-radius: 6px;
                padding: 10px;
                min-width: 80px;
            }
            
            QPushButton:hover {
                background-color: #45a049;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
            }
            
            QPushButton:pressed {
                background-color: #3d8b40;
                box-shadow: 0 1px 2px rgba(0, 0, 0, 0.3);
            }
            
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
            
            /* 输入框样式 */
            QLineEdit {
                font-size: 14px;
                font-family: "微软雅黑";
                padding: 8px;
                border: 1px solid #ddd;
                border-radius: 4px;
                background-color: white;
            }
            
            QLineEdit:focus {
                border-color: #4CAF50;
                outline: none;
            }
            
            /* 标签样式 */
            QLabel {
                font-size: 13px;
                font-family: "微软雅黑";
                color: #333;
            }
            
            /* 复选框样式 */
            QCheckBox {
                font-size: 13px;
                font-family: "微软雅黑";
                color: #333;
                padding: 3px;
            }
            
            QCheckBox:hover {
                color: #4CAF50;
            }
            
            /* 滚动区域样式 */
            QScrollArea {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                background-color: white;
            }
            
            QScrollBar:vertical {
                background: #f1f1f1;
                width: 10px;
                border-radius: 5px;
            }
            
            QScrollBar::handle:vertical {
                background: #c1c1c1;
                border-radius: 5px;
            }
            
            QScrollBar::handle:vertical:hover {
                background: #a8a8a8;
            }
        """)

    def select_data_file_path(self):
        """选择数据文件夹并显示"""
        folder = QFileDialog.getExistingDirectory(self, "选择数据文件夹", self.last_used_folder)
        if folder:
            self.file_path = folder
            self.data_file_path_edit.setText(folder)
            self.settings.setValue("last_used_folder", folder)  # 记住路径
            self.load_files(folder)
            self.update_status("selection")

    def select_save_path(self):
        """选择保存文件夹"""
        folder = QFileDialog.getExistingDirectory(self, "选择保存文件夹", self.last_output_folder)
        if folder:
            self.output_folder = folder
            self.save_path_edit.setText(folder)
            self.settings.setValue("last_output_folder", folder)  # 记住路径

    def load_files(self, folder):
        """加载文件夹中的所有文件"""
        for i in reversed(range(self.files_layout.count())):
            widget = self.files_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()

        self.selected_files.clear()
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            if os.path.isfile(file_path):
                checkbox = QCheckBox(filename, self)
                checkbox.clicked.connect(lambda: self.update_status("selection"))
                self.files_layout.addWidget(checkbox)
                self.selected_files.append((checkbox, file_path))

        self.files_group.setEnabled(True)
        self.update_status("selection")

    def refresh(self):
        """刷新文件列表"""
        if self.file_path:
            self.load_files(self.file_path)
            self.update_status("selection")

    def select_all_files(self):
        """全选所有文件"""
        for checkbox, _ in self.selected_files:
            checkbox.setChecked(True)
        self.update_status()

    def deselect_all_files(self):
        """反选所有文件"""
        for checkbox, _ in self.selected_files:
            checkbox.setChecked(not checkbox.isChecked())
        self.update_status()

    def update_status(self, status_type="selection"):
        """更新状态栏信息"""
        if status_type == "selection":
            # 显示文件选择状态
            total_files = len(self.selected_files)
            selected_count = sum(1 for checkbox, _ in self.selected_files if checkbox.isChecked())
            if total_files > 0:
                self.status_bar.showMessage(
                    f"已选择 {selected_count}/{total_files} 个文件 | 当前路径: {os.path.basename(self.file_path) if self.file_path else '未选择'}")
            else:
                self.status_bar.showMessage(
                    f"当前路径: {os.path.basename(self.file_path) if self.file_path else '未选择'} | 请选择文件夹开始")
        elif status_type == "ready":
            # 应用就绪状态
            self.status_bar.showMessage("就绪 | 请选择数据文件夹和要解析的文件")
        elif status_type == "parsing":
            # 解析中状态
            self.status_bar.showMessage("正在解析数据...")

    def start_parsing(self):
        """开始解析被选中的文件"""
        if not self.file_path:
            QMessageBox.warning(self, "警告", "请先选择数据文件夹")
            return

        if not self.output_folder:
            QMessageBox.warning(self, "警告", "请先选择保存路径")
            return

        selected_files = [file_path for checkbox, file_path in self.selected_files if checkbox.isChecked()]
        if not selected_files:
            QMessageBox.warning(self, "警告", "请至少选择一个文件进行解析")
            return

        # 开始解析每个文件并显示进度
        total_files = len(selected_files)
        for idx, file_path in enumerate(selected_files):
            self.status_bar.showMessage(f"正在解析 {idx + 1}/{total_files} 文件: {os.path.basename(file_path)}")
            structs = parse_file(file_path)
            if structs:
                timestamp_filename = generate_timestamp_filename()
                output_full_path = os.path.join(self.output_folder, timestamp_filename)
                save_to_excel(structs, output_full_path)

        # 解析成功提示
        self.status_bar.showMessage("所有文件解析完成！")
        QMessageBox.information(self, "成功", f"已成功解析 {total_files} 个文件")
        self.update_status("selection")


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
