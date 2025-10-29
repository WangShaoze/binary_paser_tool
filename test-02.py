import ctypes
import base64
import openpyxl
from openpyxl import Workbook
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime


# 定义结构体（同之前，单个结构体大小1112字节）
class STExportValue_t(ctypes.Structure):
    _fields_ = [
        ("data_len", ctypes.c_int),
        ("data_buf", ctypes.c_char * 1024),  # 512 * 2=1024
        ("time_len", ctypes.c_int),
        ("time_buf", ctypes.c_char * 32),
        ("protocol_len", ctypes.c_int),
        ("protocol_type_buf", ctypes.c_char * 7),
        ("analysis_len", ctypes.c_int),
        ("analysis_result_buf", ctypes.c_char * 32),
    ]


# 单个结构体的字节大小（验证是否为1112）
STRUCT_SIZE = ctypes.sizeof(STExportValue_t)
assert STRUCT_SIZE == 1112, f"结构体大小错误，实际为{STRUCT_SIZE}字节"


def parse_multiple_structs(binary_data):
    """解析多组结构体数据，返回列表"""
    struct_list = []
    data_length = len(binary_data)

    # 检查总长度是否为结构体大小的整数倍（避免数据不完整）
    if data_length % STRUCT_SIZE != 0:
        raise ValueError(f"二进制数据长度不正确，应为{STRUCT_SIZE}的整数倍，实际为{data_length}")

    # 按结构体大小分割并解析
    for i in range(0, data_length, STRUCT_SIZE):
        # 截取当前结构体的二进制数据
        chunk = binary_data[i:i + STRUCT_SIZE]
        # 转换为结构体实例
        struct_instance = STExportValue_t.from_buffer_copy(chunk)
        struct_list.append(struct_instance)

    return struct_list


def create_excel_template():
    """创建Excel模板，前2行保持不变"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 第一行：标题 - 合并ABCD四列
    ws.merge_cells('A1:D1')
    ws['A1'] = "馈线终端离线分析装置-导出数据表"
    
    # 设置合并单元格的样式使其居中显示
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 第二行：表头
    headers = ["时间", "协议类型", "源数据", "分析结果"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)

    return wb, ws


def generate_timestamp_filename():
    """生成包含当前日期时间的文件名"""
    now = datetime.now()
    timestamp = now.strftime("%Y年%m月%d日%H时%M分%S秒")
    return f"导出数据表-{timestamp}.xlsx"


def save_to_excel(structs, output_filename=None):
    """将解析的数据保存到Excel文件"""
    # 如果没有提供文件名，使用当前日期时间生成
    if output_filename is None:
        output_filename = generate_timestamp_filename()

    # 创建或加载模板
    if os.path.exists(output_filename):
        wb = openpyxl.load_workbook(output_filename)
        ws = wb.active
        # 清空第3行及以后的数据（保留前2行）
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)
    else:
        wb, ws = create_excel_template()

    # 从第3行开始写入数据
    start_row = 3

    for idx, uni in enumerate(structs):
        row = start_row + idx

        # 处理时间字段
        time_buf = uni.time_buf.decode('utf-8', errors='ignore')
        time_buf = " ".join([item.strip() for item in time_buf.strip().split("\n")])

        # 处理协议类型字段
        protocol_type_buf = uni.protocol_type_buf.decode('utf-8', errors='ignore')

        # 处理源数据字段（Base64解码）
        data_buf_base64_str = uni.data_buf.decode('utf-8', errors='ignore')
        try:
            decoded_bytes = base64.b64decode(data_buf_base64_str.encode('utf-8'))
            data_buf = decoded_bytes.decode('utf-8', errors='ignore')
        except Exception as e:
            data_buf = f"解码错误: {str(e)}"

        # 处理分析结果字段
        analysis_result_buf = uni.analysis_result_buf.decode('utf-8', errors='ignore')

        # 写入Excel
        ws.cell(row=row, column=1, value=time_buf)  # A列：时间
        ws.cell(row=row, column=2, value=protocol_type_buf)  # B列：协议类型
        ws.cell(row=row, column=3, value=data_buf)  # C列：源数据
        ws.cell(row=row, column=4, value=analysis_result_buf)  # D列：分析结果

    # 保存文件
    wb.save(output_filename)
    print(f"数据已成功导出到 {output_filename}")


def parse_file(file_path, output_folder):
    """解析指定的二进制文件并导出到Excel"""
    try:
        # 读取二进制文件
        with open(file_path, "rb") as f:
            all_binary_data = f.read()
        
        # 解析所有数据
        structs = parse_multiple_structs(all_binary_data)
        print(f"成功解析 {len(structs)} 组数据")
        
        # 生成时间戳文件名并拼接完整路径
        timestamp_filename = generate_timestamp_filename()
        output_full_path = os.path.join(output_folder, timestamp_filename)
        
        # 将数据保存到Excel
        save_to_excel(structs, output_full_path)
        
        # 创建自定义消息框，显示成功信息并提供复制路径和打开文件夹的选项
        show_success_dialog(len(structs), output_full_path)
        
        return structs
    except ValueError as e:
        error_msg = f"解析失败：{e}"
        print(error_msg)
        messagebox.showerror("错误", error_msg)
    except Exception as e:
        error_msg = f"处理过程中发生错误：{e}"
        print(error_msg)
        messagebox.showerror("错误", error_msg)
    return None

def show_success_dialog(data_count, output_path):
    """显示自定义成功对话框，包含复制路径和打开文件夹按钮"""
    dialog = tk.Toplevel()
    dialog.title("解析成功")
    dialog.geometry("550x200")
    dialog.resizable(False, False)
    
    # 居中显示
    dialog.update_idletasks()
    width = dialog.winfo_width()
    height = dialog.winfo_height()
    x = (dialog.winfo_screenwidth() // 2) - (width // 2)
    y = (dialog.winfo_screenheight() // 2) - (height // 2)
    dialog.geometry(f"{width}x{height}+{x}+{y}")
    
    # 成功信息
    frame = tk.Frame(dialog, padx=20, pady=15)
    frame.pack(fill=tk.BOTH, expand=True)
    
    tk.Label(frame, text="数据解析成功！", font=("SimHei", 12, "bold"), width=12).pack(pady=5)
    tk.Label(frame, text=f"共解析 {data_count} 组数据", font=("SimHei", 10)).pack(pady=2)
    
    # 文件路径显示和复制按钮
    path_frame = tk.Frame(frame)
    path_frame.pack(fill=tk.X, pady=10)
    
    path_var = tk.StringVar(value=output_path)
    path_entry = tk.Entry(path_frame, textvariable=path_var, width=50)
    path_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    
    def copy_path():
        dialog.clipboard_clear()
        dialog.clipboard_append(output_path)
        dialog.update()  # 保持剪贴板内容
        messagebox.showinfo("复制成功", "文件路径已复制到剪贴板")
    
    tk.Button(path_frame, text="复制路径", command=copy_path, width=12).pack(side=tk.RIGHT, padx=2)
    
    # 按钮区域
    button_frame = tk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=5)
    
    def open_folder():
        # 获取文件夹路径
        folder_path = os.path.dirname(output_path)
        # 在Windows下使用explorer打开文件夹
        if os.name == 'nt':  # Windows
            os.startfile(folder_path)
        else:  # macOS或Linux
            import subprocess
            subprocess.Popen(['xdg-open', folder_path])
    
    tk.Button(button_frame, text="打开文件夹", command=open_folder, width=12).pack(side=tk.LEFT, padx=2)
    tk.Button(button_frame, text="确定", command=dialog.destroy, width=12).pack(side=tk.RIGHT, padx=2)


def create_gui():
    """创建图形用户界面"""
    root = tk.Tk()
    root.title("二进制数据解析工具")
    root.geometry("600x350")
    
    # 变量存储选择的文件路径和输出文件夹路径
    file_path_var = tk.StringVar()
    output_folder_var = tk.StringVar()
    
    # 创建文件选择函数
    def select_file():
        file_path = filedialog.askopenfilename(
            title="选择二进制数据文件",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if file_path:
            file_path_var.set(file_path)
    
    # 创建文件夹选择函数
    def select_output_folder():
        folder_path = filedialog.askdirectory(
            title="选择Excel文件保存位置"
        )
        if folder_path:
            output_folder_var.set(folder_path)

    # 创建解析函数
    def start_parsing():
        file_path = file_path_var.get()
        output_folder = output_folder_var.get()
        
        # 验证文件选择
        if not file_path:
            messagebox.showwarning("警告", "请先选择一个文件")
            return
        
        # 验证文件夹选择
        if not output_folder:
            messagebox.showwarning("警告", "请先选择输出文件夹")
            return
        
        # 验证文件夹是否存在
        if not os.path.exists(output_folder):
            messagebox.showerror("错误", "选择的输出文件夹不存在")
            return
        
        # 开始解析
        parse_file(file_path, output_folder)

    # 创建界面元素
    frame = tk.Frame(root, padx=20, pady=20)
    frame.pack(fill=tk.BOTH, expand=True)
    
    # 文件选择区域
    file_frame = tk.Frame(frame)
    file_frame.pack(fill=tk.X, pady=10)
    
    tk.Label(file_frame, text="文件路径:", font=("SimHei", 10), width=12).pack(side=tk.LEFT)
    tk.Entry(file_frame, textvariable=file_path_var, width=40).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Button(file_frame, text="浏览", command=select_file).pack(side=tk.LEFT, padx=5)
    
    # 输出文件夹选择区域
    folder_frame = tk.Frame(frame)
    folder_frame.pack(fill=tk.X, pady=10)
    
    tk.Label(folder_frame, text="输出文件夹:", font=("SimHei", 10), width=12).pack(side=tk.LEFT)
    tk.Entry(folder_frame, textvariable=output_folder_var, width=40).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
    tk.Button(folder_frame, text="浏览", command=select_output_folder).pack(side=tk.LEFT, padx=5)
    
    # 解析按钮
    button_frame = tk.Frame(frame)
    button_frame.pack(fill=tk.X, pady=15)
    
    parse_button = tk.Button(
        button_frame,
        text="开始解析并导出Excel",
        command=start_parsing,
        font=("SimHei", 12),
        width=25,
        height=2
    )
    parse_button.pack(side=tk.TOP, pady=10)
    
    # 说明标签
    info_text = "使用说明：\n1. 点击'浏览'选择要解析的二进制数据文件\n2. 点击'浏览'选择Excel文件的保存位置\n3. 点击'开始解析并导出Excel'开始处理\n4. 处理完成后会显示文件保存位置，并可选择打开文件夹"
    info_label = tk.Label(frame, text=info_text, justify=tk.LEFT, font=("SimHei", 9))
    info_label.pack(fill=tk.BOTH, expand=True, pady=10)

    # 启动主循环
    root.mainloop()


# ------------------------------
# 主程序入口
# ------------------------------
if __name__ == "__main__":
    # 创建并显示GUI
    create_gui()
