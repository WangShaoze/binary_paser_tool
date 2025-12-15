import ctypes
import base64
import openpyxl
import os
from datetime import datetime
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QVBoxLayout, QHBoxLayout, QWidget, QPushButton, \
    QCheckBox, QLineEdit, QGroupBox, QDialog, QDialogButtonBox
from PyQt5.QtCore import Qt, QSettings
import sys


# 定义结构体（同之前，单个结构体大小1112字节）
class STExportValue_t(ctypes.Structure):
    _fields_ = [
        ("data_len", ctypes.c_int),
        ("data_buf", ctypes.c_char * 1024),  # 512 * 2 = 1024
        ("time_len", ctypes.c_int),
        ("time_buf", ctypes.c_char * 32),
        ("protocol_len", ctypes.c_int),
        ("protocol_type_buf", ctypes.c_char * 7),
        ("analysis_len", ctypes.c_int),
        ("analysis_result_buf", ctypes.c_char * 32),
    ]


# 单个结构体的字节大小（验证是否为1112）
STRUCT_SIZE = ctypes.sizeof(STExportValue_t)
assert STRUCT_SIZE == 1112, f"结构体大小错误，实际为{STRUCT_SIZE}字节"


def parse_multiple_structs(binary_data):
    """解析多组结构体数据，返回列表"""
    struct_list = []
    data_length = len(binary_data)

    # 检查总长度是否为结构体大小的整数倍（避免数据不完整）
    if data_length % STRUCT_SIZE != 0:
        raise ValueError(f"二进制数据长度不正确，应为{STRUCT_SIZE}的整数倍，实际为{data_length}")

    # 按结构体大小分割并解析
    for i in range(0, data_length, STRUCT_SIZE):
        # 截取当前结构体的二进制数据
        chunk = binary_data[i:i + STRUCT_SIZE]
        # 转换为结构体实例
        struct_instance = STExportValue_t.from_buffer_copy(chunk)
        struct_list.append(struct_instance)

    return struct_list


def create_excel_template():
    """创建Excel模板，前2行保持不变"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 第一行：标题 - 合并ABCD四列
    ws.merge_cells('A1:D1')
    ws['A1'] = "馈线终端离线分析装置-导出数据表"

    # 设置合并单元格的样式使其居中显示
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    # 第二行：表头
    headers = ["时间", "协议类型", "源数据", "分析结果"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=header)

    return wb, ws


def generate_timestamp_filename():
    """生成包含当前日期时间的文件名"""
    now = datetime.now()
    timestamp = now.strftime("%Y年%m月%d日%H时%M分%S秒")
    return f"导出数据表-{timestamp}.xlsx"


def save_to_excel(structs, output_filename=None):
    """将解析的数据保存到Excel文件"""
    # 如果没有提供文件名，使用当前日期时间生成
    if output_filename is None:
        output_filename = generate_timestamp_filename()

    # 如果文件已存在，则加载文件
    if os.path.exists(output_filename):
        wb = openpyxl.load_workbook(output_filename)
        ws = wb.active
        # 找到最后一行的下一行作为开始写入的行
        start_row = ws.max_row + 1
    else:
        # 如果文件不存在，则创建一个新的文件并写入表头
        wb, ws = create_excel_template()
        start_row = 3  # 从第三行开始写入数据，因为前两行是表头

    # 从指定的起始行开始写入数据
    for idx, uni in enumerate(structs):
        row = start_row + idx

        # 处理时间字段
        time_buf = uni.time_buf.decode('utf-8', errors='ignore')
        time_buf = " ".join([item.strip() for item in time_buf.strip().split("\n")])

        # 处理协议类型字段
        protocol_type_buf = uni.protocol_type_buf.decode('utf-8', errors='ignore')

        # 处理源数据字段（Base64解码）
        data_buf_base64_str = uni.data_buf.decode('utf-8', errors='ignore')
        try:
            decoded_bytes = base64.b64decode(data_buf_base64_str.encode('utf-8'))
            data_buf = decoded_bytes.decode('utf-8', errors='ignore')
        except Exception as e:
            data_buf = f"解码错误: {str(e)}"

        # 处理分析结果字段
        analysis_result_buf = uni.analysis_result_buf.decode('utf-8', errors='ignore')

        # 写入Excel
        ws.cell(row=row, column=1, value=time_buf)  # A列：时间
        ws.cell(row=row, column=2, value=protocol_type_buf)  # B列：协议类型
        ws.cell(row=row, column=3, value=data_buf)  # C列：源数据
        ws.cell(row=row, column=4, value=analysis_result_buf)  # D列：分析结果

    # 保存文件
    wb.save(output_filename)
    print(f"数据已成功导出到 {output_filename}")


def parse_file(file_path):
    """解析指定的二进制文件并导出到Excel"""
    try:
        # 读取二进制文件
        with open(file_path, "rb") as f:
            all_binary_data = f.read()

        # 解析所有数据
        structs = parse_multiple_structs(all_binary_data)
        print(f"成功解析 {len(structs)} 组数据")

        return structs
    except ValueError as e:
        print(f"解析失败：{e}")
    except Exception as e:
        print(f"处理过程中发生错误：{e}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("二进制数据解析工具")
        self.setGeometry(100, 100, 800, 600)

        self.file_path = None
        self.selected_files = []
        self.output_folder = None

        # 加载上次选择的路径
        self.settings = QSettings("MyCompany", "BinaryParserApp")
        self.last_used_folder = self.settings.value("last_used_folder", "", type=str)

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # 文件路径选择区域
        file_group = QGroupBox("选择文件路径")
        file_layout = QHBoxLayout()
        self.file_line_edit = QLineEdit(self)
        self.file_line_edit.setReadOnly(True)
        self.file_line_edit.setFixedHeight(30)  # 调整文件路径框的高度
        file_button = QPushButton("浏览", self)
        file_button.clicked.connect(self.select_file)

        refresh_button = QPushButton("刷新", self)  # 放置刷新按钮
        refresh_button.clicked.connect(self.refresh)

        file_layout.addWidget(self.file_line_edit)
        file_layout.addWidget(file_button)
        file_layout.addWidget(refresh_button)  # 刷新按钮紧随浏览按钮
        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # 选择的文件区域
        self.files_group = QGroupBox("选择需要解析的文件")
        self.files_group.setEnabled(False)  # 默认禁用
        self.files_layout = QVBoxLayout()
        self.files_group.setLayout(self.files_layout)
        layout.addWidget(self.files_group)

        # 开始解析按钮
        parse_button = QPushButton("开始解析", self)
        parse_button.clicked.connect(self.start_parsing)
        layout.addWidget(parse_button)

        # 设置主界面布局
        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

        # 设置样式
        self.set_style()

    def set_style(self):
        """设置界面样式"""
        self.setStyleSheet("""
            QGroupBox {
                font: bold 12pt;
                border: 1px solid gray;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 10px;
            }
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                border-radius: 5px;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QLineEdit {
                font-size: 14px;
                padding: 5px;
            }
            QLabel {
                font-size: 12px;
            }
        """)

    def select_file(self):
        """选择文件夹并显示文件"""
        folder = QFileDialog.getExistingDirectory(self, "选择文件夹", self.last_used_folder)
        if folder:
            self.file_path = folder
            self.file_line_edit.setText(folder)
            self.settings.setValue("last_used_folder", folder)  # 记住路径
            self.load_files(folder)

    def load_files(self, folder):
        """加载文件夹中的所有文件"""
        for i in reversed(range(self.files_layout.count())):
            widget = self.files_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()

        self.selected_files.clear()
        for filename in os.listdir(folder):
            file_path = os.path.join(folder, filename)
            if os.path.isfile(file_path):
                checkbox = QCheckBox(filename, self)
                self.files_layout.addWidget(checkbox)
                self.selected_files.append((checkbox, file_path))

        self.files_group.setEnabled(True)

    def refresh(self):
        """刷新文件列表"""
        if self.file_path:
            self.load_files(self.file_path)

    def start_parsing(self):
        """开始解析被选中的文件"""
        if not self.file_path:
            print("请选择文件夹")
            return

        selected_files = [file_path for checkbox, file_path in self.selected_files if checkbox.isChecked()]
        if not selected_files:
            print("请选择至少一个文件进行解析")
            return

        output_folder = QFileDialog.getExistingDirectory(self, "选择保存文件夹")
        if not output_folder:
            print("请选择输出文件夹")
            return

        # 开始解析每个文件
        for file_path in selected_files:
            print(f"解析文件: {file_path}")
            structs = parse_file(file_path)
            if structs:
                timestamp_filename = generate_timestamp_filename()
                output_full_path = os.path.join(output_folder, timestamp_filename)
                save_to_excel(structs, output_full_path)
                print(f"数据已保存到: {output_full_path}")


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
